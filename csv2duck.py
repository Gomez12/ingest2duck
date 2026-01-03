# csv2duck.py  (library; no CLI)  -> supports CSV + XLSX with raw + normalized
from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from typing import Any, Dict, Iterator, List, Optional

import dlt
from openpyxl import load_workbook

from utils import sanitize_table_name, get_pk_from_record, stable_pk_from_fields


@dataclass
class TabularCollection:
    name: str
    kind: str
    sheet: Optional[str] = None
    use_first_sheet: bool = False
    pk_prefer: Optional[List[str]] = None
    write_disposition: Optional[str] = None
    enabled: bool = True


@dataclass
class TabularMapping:
    collections: Dict[str, TabularCollection]


def _infer_pk_prefer_from_header(header: List[str]) -> List[str]:
    base = ["id", "ID", "code", "Code", "key", "Key", "pk", "PK"]
    lower = {h.lower(): h for h in header}
    out: List[str] = []
    for b in base:
        if b in header:
            out.append(b)
        else:
            hit = lower.get(b.lower())
            if hit:
                out.append(hit)
    # unique
    uniq: List[str] = []
    for x in out:
        if x not in uniq:
            uniq.append(x)
    return uniq


def iter_csv_rows(path: str, delimiter: str = ",", encoding: str = "utf-8") -> Iterator[Dict[str, Any]]:
    with open(path, "r", encoding=encoding, newline="") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        for row in reader:
            out: Dict[str, Any] = {}
            for k, v in (row or {}).items():
                kk = (k or "").strip()
                if kk == "":
                    continue
                vv = v.strip() if isinstance(v, str) else v
                out[kk] = vv if (vv is not None and str(vv).strip() != "") else None
            yield out


def iter_xlsx_rows(path: str, sheet: Optional[str] = None, use_first_sheet: bool = False) -> Iterator[Dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if use_first_sheet:
            ws = wb.worksheets[0]
        elif sheet:
            ws = wb[sheet]
        else:
            ws = wb.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        header: Optional[List[str]] = None

        for r in rows:
            if r is None:
                continue
            if header is None:
                hdr = [str(x).strip() if x is not None and str(x).strip() != "" else "" for x in r]
                if any(hdr):
                    header = hdr
                continue

            d: Dict[str, Any] = {}
            for i, col in enumerate(header):
                if not col:
                    continue
                val = r[i] if i < len(r) else None
                # keep scalars; stringify everything else
                if val is None:
                    d[col] = None
                elif isinstance(val, (str, int, float, bool)):
                    d[col] = val
                else:
                    try:
                        import datetime
                        if isinstance(val, (datetime.datetime, datetime.date)):
                            d[col] = val.isoformat()
                        else:
                            d[col] = str(val)
                    except Exception:
                        d[col] = str(val)

            if any(v is not None and str(v).strip() != "" for v in d.values()):
                yield d
    finally:
        try:
            wb.close()
        except Exception:
            pass


def infer_tabular_mapping_for_csv(source_name: str, table: str, sample_path: str, delimiter: str = ",", encoding: str = "utf-8") -> TabularMapping:
    # read header via DictReader fieldnames
    with open(sample_path, "r", encoding=encoding, newline="") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        header = reader.fieldnames or []
    pk_prefer = _infer_pk_prefer_from_header([h for h in header if h])
    collection_name = f"{source_name}_{table}"
    col = TabularCollection(name=collection_name, kind="csv", sheet=None, pk_prefer=pk_prefer)
    return TabularMapping(collections={collection_name: col})


def infer_tabular_mapping_for_xlsx(source_name: str, table: str, sample_path: str, sheet: Optional[str] = None, use_first_sheet: bool = False) -> TabularMapping:
    wb = load_workbook(sample_path, read_only=True, data_only=True)
    try:
        if use_first_sheet:
            ws = wb.worksheets[0]
        elif sheet:
            ws = wb[sheet]
        else:
            ws = wb.worksheets[0]
        header: List[str] = []
        for r in ws.iter_rows(values_only=True):
            if r is None:
                continue
            hdr = [str(x).strip() if x is not None and str(x).strip() != "" else "" for x in r]
            if any(hdr):
                header = [h for h in hdr if h]
                break
        pk_prefer = _infer_pk_prefer_from_header(header)
        sheet_name = sheet or ws.title
        collection_name = f"{source_name}_{table}"
        col = TabularCollection(name=collection_name, kind="xlsx", sheet=sheet_name, use_first_sheet=use_first_sheet, pk_prefer=pk_prefer)
        return TabularMapping(collections={collection_name: col})
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _create_normalized_resource(
    tname: str,
    col: TabularCollection,
    file_path: str,
    fmt: str,
    delimiter: str,
    encoding: str,
    write_disposition: str,
    pk_fields_override: Optional[List[str]],
    source_name: str,
) -> Any:
    def _iter_rows() -> Iterator[Dict[str, Any]]:
        if fmt == "csv":
            yield from iter_csv_rows(file_path, delimiter=delimiter, encoding=encoding)
        elif fmt == "xlsx":
            yield from iter_xlsx_rows(file_path, sheet=col.sheet, use_first_sheet=col.use_first_sheet)
        else:
            raise ValueError(f"Unsupported tabular fmt: {fmt}")

    @dlt.resource(name=tname, write_disposition=write_disposition)
    def _resource():
        for row in _iter_rows():
            if pk_fields_override:
                pk = stable_pk_from_fields(row, pk_fields_override)
            else:
                pk = get_pk_from_record(row, col.pk_prefer or []) or stable_pk_from_fields(row, None)

            out = {"_pk": str(pk), "raw_json": json.dumps(row, ensure_ascii=False), "__source_name": source_name}
            if fmt == "xlsx" and col.sheet:
                out["__xlsx_sheet"] = col.sheet
            out.update(row)
            yield out

    return _resource


def build_tabular_resources(
    file_path: str,
    fmt: str,
    mapping: TabularMapping,
    source_name: str,
    raw_table: str,
    write_disposition: str = "append",
    delimiter: str = ",",
    encoding: str = "utf-8",
    pk_fields_override: Optional[List[str]] = None,
):
    raw_table = sanitize_table_name(raw_table, "raw_ingest")

    def _iter_rows(col: TabularCollection) -> Iterator[Dict[str, Any]]:
        if fmt == "csv":
            yield from iter_csv_rows(file_path, delimiter=delimiter, encoding=encoding)
        elif fmt == "xlsx":
            yield from iter_xlsx_rows(file_path, sheet=col.sheet, use_first_sheet=col.use_first_sheet)
        else:
            raise ValueError(f"Unsupported tabular fmt: {fmt}")

    @dlt.resource(name=raw_table, write_disposition=write_disposition)
    def raw_resource():
        for cname, col in mapping.collections.items():
            if not col.enabled:
                continue
            for row in _iter_rows(col):
                yield {
                    "collection": cname,
                    "path": (col.sheet or ""),
                    "__source_name": source_name,
                    "raw_json": json.dumps(row, ensure_ascii=False),
                }

    normalized_resources: Dict[str, Any] = {}

    for cname, col in mapping.collections.items():
        if not col.enabled:
            continue
        tname = sanitize_table_name(cname, "tabular_data")
        col_write_disposition = col.write_disposition or write_disposition
        normalized_resources[cname] = _create_normalized_resource(
            tname=tname,
            col=col,
            file_path=file_path,
            fmt=fmt,
            delimiter=delimiter,
            encoding=encoding,
            write_disposition=col_write_disposition,
            pk_fields_override=pk_fields_override,
            source_name=source_name,
        )

    return raw_resource, normalized_resources
